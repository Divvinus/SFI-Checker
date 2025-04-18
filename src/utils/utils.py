import asyncio
import random

from eth_account import Account

from src.logger import AsyncLogger

Account.enable_unaudited_hdwallet_features()

_ACCOUNT = Account()

def get_address(mnemonic: str) -> str:
    normalized_mnemonic = ' '.join(word for word in mnemonic.split() if word)
    
    if len(normalized_mnemonic.split()) in (12, 24):
        return _ACCOUNT.from_mnemonic(normalized_mnemonic).address
    else:
        if not mnemonic.startswith('0x'):
            mnemonic = '0x' + mnemonic
        return _ACCOUNT.from_key(mnemonic).address

async def random_sleep(
    address: str | None = None, 
    min_sec: int = 30, 
    max_sec: int = 60
) -> None:
    logger = AsyncLogger()
    delay = random.uniform(min_sec, max_sec)
    
    minutes, seconds = divmod(delay, 60)
    template = (
        f"Sleep "
        f"{int(minutes)} minutes {seconds:.1f} seconds" if minutes > 0 else 
        f"Sleep {seconds:.1f} seconds"
    )
    await logger.logger_msg(template, type_msg="info", address=address)
    
    chunk_size = 0.1
    chunks = int(delay / chunk_size)
    remainder = delay - (chunks * chunk_size)
    
    try:
        for _ in range(chunks):
            await asyncio.sleep(chunk_size)
            
        if remainder > 0:
            await asyncio.sleep(remainder)
            
    except asyncio.CancelledError:
        await logger.logger_msg(
            f"Sleep interrupted", type_msg="warning", address=address
        )
        raise

async def update_token_balance(account: "Account", token_amount: str | int | float) -> bool:
    from pathlib import Path
    import openpyxl
    from src.logger import AsyncLogger
    
    logger = AsyncLogger()
    base_path = Path(__file__).parent.parent.parent
    accounts_path = base_path / 'config' / 'data' / 'client' / 'accounts.xlsx'
    
    wallet_address = get_address(account.mnemonic)
    
    try:
        wb = openpyxl.load_workbook(accounts_path)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        mnemonic_idx = None
        tokens_idx = None
        
        for idx, header in enumerate(headers):
            if header == "Mnemonic":
                mnemonic_idx = idx
            elif header == "Tokens":
                tokens_idx = idx
        
        if mnemonic_idx is None:
            await logger.logger_msg(
                "Column 'Mnemonic' not found in accounts file",
                type_msg="error", address=wallet_address
            )
            return False
        
        if tokens_idx is None:
            tokens_idx = len(headers)
            ws.cell(row=1, column=tokens_idx + 1, value="Tokens")
        
        target_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if row[mnemonic_idx] == account.mnemonic:
                target_row = row_idx
                break
        
        if target_row is None:
            await logger.logger_msg(
                "Account not found in file",
                type_msg="warning", address=wallet_address
            )
            return False
        
        ws.cell(row=target_row, column=tokens_idx + 1, value=token_amount)
        wb.save(accounts_path)
        
        await logger.logger_msg(
            f"Tokens ({token_amount}) successfully written",
            type_msg="success", address=wallet_address
        )
        return True
        
    except Exception as e:
        await logger.logger_msg(
            f"Error updating tokens: {str(e)}",
            type_msg="error", address=wallet_address
        )
        return False
